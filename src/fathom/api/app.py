"""FastAPI app exposing the Architect.IQ engine (spec §5).

Endpoints back the React UI: create/list/get/edit estimates, list patterns,
extract dropped context files, recompute under deal-shaping knobs, and record
actuals for calibration.
"""

from __future__ import annotations

import io
import os
import re

import jwt
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

from .. import data_loader
from ..auth import oidc
from ..auth.access import resolve_access, visible_estimate_filter
from ..auth.security import create_access_token, decode_access_token, verify_password
from ..core import estimation
from ..core.recompute import RecomputeOverrides, recompute
from ..emit.mermaid import architecture_mermaid
from ..integrations.notion import get_opportunity_notes
from ..memory.priors import ActualOutcome
from ..memory.retrieval import Reference
from ..models.context import ContextPanel
from ..models.enums import Location
from ..models.org import AccessContext, Permission, Role, User
from ..models.results import ClientContext
from ..models.scenario import Scenario
from ..models.solution_graph import SolutionGraph
from ..models.team import RateRow
from ..models.variables import Variables
from ..service import EstimateService

app = FastAPI(title="Architect.IQ", version="0.1.0")

# The Vite dev server origin; CORS_ORIGINS overrides in other environments.
_origins = os.environ.get("FATHOM_CORS", "http://localhost:5173").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

service = EstimateService(db_path=os.environ.get("FATHOM_DB", "fathom.db"))


# --- Auth dependencies ---

def get_current_user(authorization: str | None = Header(default=None)) -> User:
    """Resolve the bearer token to a user, or 401."""
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="missing bearer token")
    token = authorization.split(" ", 1)[1]
    try:
        claims = decode_access_token(token)
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="invalid or expired token")
    user = service.directory.get_user(claims.get("sub", ""))
    if user is None:
        raise HTTPException(status_code=401, detail="user not found")
    return user


def require_admin(user: User = Depends(get_current_user)) -> User:
    if user.role is not Role.ADMIN:
        raise HTTPException(status_code=403, detail="admin only")
    return user


def _access_or_403(user: User, estimate_id: str, need: str = "view") -> AccessContext:
    ctx = resolve_access(user, estimate_id, service.repo, service.directory)
    ok = {"view": ctx.can_view, "comment": ctx.can_comment, "edit": ctx.can_edit}[need]
    if not ok:
        raise HTTPException(status_code=403, detail=f"{need} not permitted on this estimate")
    return ctx


# --- Auth endpoints ---

class LoginRequest(BaseModel):
    email: str
    password: str


class TokenResponse(BaseModel):
    token: str
    user: User


@app.post("/api/auth/login", response_model=TokenResponse)
def login(req: LoginRequest) -> TokenResponse:
    user = service.directory.get_user_by_email(req.email)
    stored_hash = service.directory.password_hash_for(req.email)
    if user is None or not stored_hash or not verify_password(req.password, stored_hash):
        raise HTTPException(status_code=401, detail="invalid email or password")
    token = create_access_token(user_id=user.id, email=user.email, role=user.role.value)
    return TokenResponse(token=token, user=user)


@app.get("/api/auth/me", response_model=User)
def me(user: User = Depends(get_current_user)) -> User:
    return user


@app.get("/api/auth/providers")
def auth_providers() -> dict:
    """Which SSO providers are configured (drives the login screen)."""
    return {"local": True, "google": oidc.google_enabled()}


@app.get("/api/auth/google/login")
def google_login() -> dict:
    if not oidc.google_enabled():
        raise HTTPException(status_code=400, detail="Google SSO is not configured")
    return {"url": oidc.google_auth_url(state="fathom")}


@app.get("/api/auth/google/callback")
def google_callback(code: str) -> TokenResponse:  # pragma: no cover - needs live creds
    if not oidc.google_enabled():
        raise HTTPException(status_code=400, detail="Google SSO is not configured")
    profile = oidc.exchange_google_code(code)
    user = service.directory.upsert_oauth_user(email=profile["email"], name=profile["name"], provider="google")
    token = create_access_token(user_id=user.id, email=user.email, role=user.role.value)
    return TokenResponse(token=token, user=user)


# --- Request/response models ---

class CreateEstimateRequest(BaseModel):
    project_name: str
    prd_text: str
    client_context: ClientContext = Field(default_factory=ClientContext)
    match_override: str | None = None


class EstimateResponse(BaseModel):
    estimate_id: str
    version: int
    graph: SolutionGraph
    mermaid: str
    references: list[Reference] = []
    opportunity_id: str | None = None


class ContextExtractResponse(BaseModel):
    filename: str
    text: str


class DecomposeRequirementsRequest(BaseModel):
    text: str
    existing: list[str] = Field(default_factory=list)


class DecomposedRequirement(BaseModel):
    text: str
    kind: str
    confidence: float
    duplicate_of: str | None = None
    taxonomy_kind: str
    taxonomy_confidence: float
    title: str
    epic: str
    priority: str | None = None


class SummarizeRequest(BaseModel):
    text: str


def _to_response(stored, references: list[Reference] | None = None) -> EstimateResponse:
    return EstimateResponse(
        estimate_id=stored.estimate_id,
        version=stored.version,
        graph=stored.graph,
        mermaid=architecture_mermaid(stored.graph),
        references=references or [],
        opportunity_id=stored.opportunity_id,
    )


# --- Endpoints ---

@app.get("/api/health")
def health() -> dict:
    return {"status": "ok", "version": app.version}


@app.get("/api/demo/status")
def demo_status() -> dict:
    """Whether demo data has been seeded, and how many estimates exist."""
    from ..demo import DEMO_SCENARIOS, is_seeded

    return {
        "seeded": is_seeded(service),
        "count": len(service.list_estimates()),
        "available": len(DEMO_SCENARIOS),
    }


@app.post("/api/demo/seed")
def demo_seed(admin: User = Depends(require_admin)) -> dict:
    """Populate the store with curated demo estimates (idempotent by name)."""
    from ..demo import seed_demo

    return seed_demo(service)


@app.get("/api/patterns")
def list_patterns() -> list[dict]:
    patterns, version = data_loader.load_patterns()
    return [
        {
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "when_to_use": p.when_to_use,
            "match_signals": p.match_signals,
            "library_version": version,
        }
        for p in patterns.values()
    ]


@app.post("/api/context/extract", response_model=ContextExtractResponse)
async def extract_context(file: UploadFile = File(...), user: User = Depends(get_current_user)) -> ContextExtractResponse:
    """Extract text from a dropped context file for the UI.

    Supported: .docx, .xlsx, .csv, .pdf, images (.png/.jpg/.jpeg/.gif/.webp), and
    plain text (.md/.txt). Spreadsheets flatten to readable rows; PDFs extract
    text; images are transcribed via Claude vision (requirements or architecture
    diagrams) when an API key is set.
    """
    raw = await file.read()
    name = file.filename or "upload"
    lower = name.lower()
    try:
        if lower.endswith(".docx"):
            text = _extract_docx(raw)
        elif lower.endswith(".xlsx"):
            text = _extract_xlsx(raw)
        elif lower.endswith(".csv"):
            text = _extract_csv(raw)
        elif lower.endswith(".pdf"):
            text = _extract_pdf(raw)
        elif _image_media_type(lower):
            from ..core.vision import describe_image

            text = describe_image(raw, _image_media_type(lower))
        else:
            text = raw.decode("utf-8", errors="replace")
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=422, detail=f"could not parse {name}: {exc}")
    return ContextExtractResponse(filename=name, text=text)


@app.post("/api/context/decompose-requirements")
def decompose_requirements(req: DecomposeRequirementsRequest, user: User = Depends(get_current_user)) -> list[DecomposedRequirement]:
    """Break dropped/fetched text into atomic requirements, flagging near-duplicates
    and classifying each against the estimate-kind taxonomy.

    Used when new context lands on the Requirements tab: decompose the document
    into distinct requirement statements, checking each against `existing` (the
    current Requirements list). A substantive duplicate is still returned, with
    `duplicate_of` set to the matching existing text, so the caller can group it
    with its match and tag it rather than silently add or drop it. Each item also
    carries `taxonomy_kind`/`taxonomy_confidence` (D25's classify_kind), a short
    `title` distinct from the full `text` (so a caller can show the title and
    keep the full text as notes — detail isn't lost just because the grid shows
    a terse label), and an `epic` grouping items by what they're about (never by
    source filename) so the caller can route a genuine epic/feature/story
    straight into the estimate's work breakdown instead of adding it as a plain
    requirement entry.
    """
    from ..core import llm

    if not req.text.strip():
        return []
    if llm.available():
        try:
            items = llm.extract_new_requirements(req.text, req.existing)
        except Exception:  # noqa: BLE001 - fall back to the heuristic on any LLM error
            items = llm.heuristic_new_requirements(req.text, req.existing)
    else:
        items = llm.heuristic_new_requirements(req.text, req.existing)

    taxonomy, _ = data_loader.load_estimate_kinds()
    for it in items:
        try:
            classified = llm.classify_kind(it["text"], taxonomy) if llm.available() else llm.heuristic_classify_kind(it["text"], taxonomy)
        except Exception:  # noqa: BLE001 - fall back to the heuristic on any LLM error
            classified = llm.heuristic_classify_kind(it["text"], taxonomy)
        it["taxonomy_kind"] = classified["kind"]
        it["taxonomy_confidence"] = classified["confidence"]
        it["title"] = classified["title"]
        it["priority"] = llm.detect_priority(it["text"])

    texts = [it["text"] for it in items]
    try:
        epics = llm.group_into_epics(texts) if llm.available() else llm.heuristic_group_into_epics(texts)
    except Exception:  # noqa: BLE001 - fall back to the heuristic on any LLM error
        epics = llm.heuristic_group_into_epics(texts)
    for it, epic in zip(items, epics):
        it["epic"] = epic

    return [DecomposedRequirement(**it) for it in items]


@app.post("/api/context/summarize")
def summarize_document(req: SummarizeRequest, user: User = Depends(get_current_user)) -> dict:
    """A short 1-2 sentence summary of dropped/fetched text, for the Requirements-tab
    entry representing the source file/URL itself (distinct from the atomic
    requirement entries `decompose-requirements` extracts from the same text)."""
    from ..core import llm

    if not req.text.strip():
        return {"summary": ""}
    if llm.available():
        try:
            return {"summary": llm.summarize_document(req.text)}
        except Exception:  # noqa: BLE001 - fall back to the heuristic on any LLM error
            pass
    return {"summary": llm.heuristic_summarize(req.text)}


_IMAGE_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
}


def _image_media_type(lower_name: str) -> str | None:
    for ext, media in _IMAGE_TYPES.items():
        if lower_name.endswith(ext):
            return media
    return None


def _extract_pdf(raw: bytes, max_pages: int = 100) -> str:
    """Extract text from a PDF (requirements docs). Falls back to a note if the
    PDF is scanned/image-only with no extractable text (use image drop instead)."""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    pages = []
    for i, page in enumerate(reader.pages):
        if i >= max_pages:
            pages.append(f"... (truncated at {max_pages} pages)")
            break
        pages.append(page.extract_text() or "")
    text = "\n".join(p.strip() for p in pages if p.strip())
    if not text.strip():
        return (
            "> PDF has no extractable text (likely scanned). Drop it as an image "
            "to use vision extraction instead.\n"
        )
    return text


def _extract_docx(raw: bytes) -> str:
    import docx  # python-docx

    document = docx.Document(io.BytesIO(raw))
    return "\n".join(p.text for p in document.paragraphs)


def _extract_xlsx(raw: bytes, max_rows_per_sheet: int = 500) -> str:
    """Flatten an .xlsx to text: one section per sheet, cells joined by ' | '."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    blocks: list[str] = []
    for ws in wb.worksheets:
        lines = [f"## Sheet: {ws.title}"]
        truncated = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                truncated = True
                break
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
        if truncated:
            lines.append(f"... (truncated at {max_rows_per_sheet} rows)")
        if len(lines) > 1:
            blocks.append("\n".join(lines))
    wb.close()
    return "\n\n".join(blocks)


def _extract_csv(raw: bytes, max_rows: int = 2000) -> str:
    """Normalize a .csv to ' | '-joined rows (handles quoting/delimiters)."""
    import csv

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines: list[str] = []
    for i, row in enumerate(reader):
        if i >= max_rows:
            lines.append(f"... (truncated at {max_rows} rows)")
            break
        cells = [c.strip() for c in row if c.strip()]
        if cells:
            lines.append(" | ".join(cells))
    return "\n".join(lines)


class CreateEstimateRequestAuthed(CreateEstimateRequest):
    opportunity_id: str | None = None


@app.post("/api/estimates", response_model=EstimateResponse)
def create_estimate(req: CreateEstimateRequestAuthed, user: User = Depends(get_current_user)) -> EstimateResponse:
    if user.role is Role.CLIENT:
        raise HTTPException(status_code=403, detail="clients cannot create estimates")
    if not req.prd_text.strip():
        raise HTTPException(status_code=422, detail="prd_text is required")
    stored, references = service.create_estimate(
        req.project_name or "Untitled", req.prd_text, req.client_context, req.match_override,
        owner_id=user.id, opportunity_id=req.opportunity_id,
    )
    return _to_response(stored, references)


@app.get("/api/estimates")
def list_estimates(user: User = Depends(get_current_user)) -> list[dict]:
    """Role-scoped: admin=all, user=own+shared, client=assigned opportunities."""
    predicate = visible_estimate_filter(user, service.repo, service.directory)
    return [s.__dict__ for s in service.list_estimates() if predicate(s)]


@app.get("/api/estimates/{estimate_id}", response_model=EstimateResponse)
def get_estimate(estimate_id: str, version: int | None = None, user: User = Depends(get_current_user)) -> EstimateResponse:
    _access_or_403(user, estimate_id, "view")
    stored = service.get_estimate(estimate_id, version)
    if stored is None:
        raise HTTPException(status_code=404, detail="estimate not found")
    return _to_response(stored, service.get_references(estimate_id, version))


@app.post("/api/estimates/{estimate_id}/recompute", response_model=EstimateResponse)
def recompute_estimate(estimate_id: str, overrides: RecomputeOverrides, user: User = Depends(get_current_user)) -> EstimateResponse:
    """Apply deal-shaping knobs and persist the result as a new version (§5.5)."""
    _access_or_403(user, estimate_id, "edit")
    stored = service.get_estimate(estimate_id)
    if stored is None:
        raise HTTPException(status_code=404, detail="estimate not found")
    updated_graph = recompute(stored.graph, overrides)
    saved = service.update_estimate(estimate_id, updated_graph)
    return _to_response(saved, service.get_references(estimate_id))


@app.post("/api/estimates/{estimate_id}/rebuild", response_model=EstimateResponse)
def rebuild_estimate(estimate_id: str, req: CreateEstimateRequestAuthed, user: User = Depends(get_current_user)) -> EstimateResponse:
    """Auto-save: re-derive the estimate from updated inputs, in place (no new version)."""
    _access_or_403(user, estimate_id, "edit")
    if not req.prd_text.strip():
        raise HTTPException(status_code=422, detail="prd_text is required")
    try:
        stored, references = service.rebuild_estimate(
            estimate_id, req.project_name or "Untitled", req.prd_text, req.client_context, req.opportunity_id
        )
    except KeyError:
        raise HTTPException(status_code=404, detail="estimate not found")
    return _to_response(stored, references)


class TagsRequest(BaseModel):
    tags: list[str]


@app.post("/api/estimates/{estimate_id}/tags", response_model=EstimateResponse)
def set_tags(estimate_id: str, req: TagsRequest, user: User = Depends(get_current_user)) -> EstimateResponse:
    """Set estimate metadata tags (auto-saved in place, no new version)."""
    _access_or_403(user, estimate_id, "edit")
    stored = service.get_estimate(estimate_id)
    if stored is None:
        raise HTTPException(status_code=404, detail="estimate not found")
    graph = stored.graph.model_copy(deep=True)
    graph.tags = [t.strip() for t in req.tags if t.strip()][:20]
    saved = service.repo.overwrite_latest(estimate_id, graph)
    return _to_response(saved, service.get_references(estimate_id))


@app.post("/api/estimates/{estimate_id}/clone", response_model=EstimateResponse)
def clone_estimate(estimate_id: str, user: User = Depends(get_current_user)) -> EstimateResponse:
    """Clone an estimate to test other assumptions (new estimate, never active)."""
    _access_or_403(user, estimate_id, "view")
    if user.role is Role.CLIENT:
        raise HTTPException(status_code=403, detail="clients cannot clone estimates")
    try:
        stored = service.clone_estimate(estimate_id, owner_id=user.id)
    except KeyError:
        raise HTTPException(status_code=404, detail="estimate not found")
    return _to_response(stored, service.get_references(stored.estimate_id))


@app.put("/api/estimates/{estimate_id}", response_model=EstimateResponse)
def update_estimate(estimate_id: str, graph: SolutionGraph, user: User = Depends(get_current_user)) -> EstimateResponse:
    """Persist a fully edited graph as a new version (interactive editing)."""
    _access_or_403(user, estimate_id, "edit")
    if service.get_estimate(estimate_id) is None:
        raise HTTPException(status_code=404, detail="estimate not found")
    saved = service.update_estimate(estimate_id, graph)
    return _to_response(saved, service.get_references(estimate_id))


@app.get("/api/rates")
def get_rates(user: User = Depends(get_current_user)) -> dict:
    """The active rate card: rows + source + summary (leverage modeling, §2.6)."""
    from ..core.rates import RateCard

    rows, source = service.active_rates()
    practices, _ = data_loader.load_practices()
    return {
        "source": source,
        "summary": RateCard(rows).summary(),
        "practices": sorted({p for r in rows if (p := practices.practice_of(r.discipline))}),
        "rates": [
            {"discipline": r.discipline, "tier": r.tier, "location": r.location.value, "day_rate": r.day_rate}
            for r in rows
        ],
    }


def _card_dict(card, include_rows: bool = False) -> dict:
    from ..core.rates import RateCard

    out = {
        "id": card.id,
        "name": card.name,
        "is_default": card.is_default,
        "is_active": card.is_active,
        "summary": RateCard(card.rows).summary(),
    }
    if include_rows:
        out["rates"] = [
            {"discipline": r.discipline, "tier": r.tier, "location": r.location.value, "day_rate": r.day_rate}
            for r in card.rows
        ]
    return out


@app.get("/api/rate-cards")
def list_rate_cards(user: User = Depends(get_current_user)) -> list[dict]:
    """All saved rate cards (one active, one default)."""
    return [_card_dict(c) for c in service.list_rate_cards()]


@app.post("/api/rate-cards")
async def create_rate_card(file: UploadFile = File(...), name: str = Form(None), admin: User = Depends(require_admin)) -> dict:
    """Save a rate card from a file (.csv/.xlsx/.yaml) and make it active."""
    from ..core.rates import parse_rate_file

    raw = await file.read()
    try:
        rows = parse_rate_file(raw, file.filename or "rates")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    card = service.create_rate_card(name or (file.filename or "Rate card"), rows)
    return _card_dict(card, include_rows=True)


class RateRowIn(BaseModel):
    discipline: str
    tier: str
    location: Location
    day_rate: float = Field(ge=0.0)


class UpdateRateCardRequest(BaseModel):
    rows: list[RateRowIn]


@app.put("/api/rate-cards/{card_id}")
def update_rate_card(card_id: str, req: UpdateRateCardRequest, admin: User = Depends(require_admin)) -> dict:
    """Replace a saved rate card's rows in place (manual edit, not a re-upload)."""
    rows = [RateRow(discipline=r.discipline, tier=r.tier, location=r.location, day_rate=r.day_rate) for r in req.rows]
    try:
        return _card_dict(service.update_rate_card(card_id, rows), include_rows=True)
    except KeyError:
        raise HTTPException(status_code=404, detail="rate card not found")


@app.post("/api/rate-cards/{card_id}/activate")
def activate_rate_card(card_id: str, admin: User = Depends(require_admin)) -> dict:
    try:
        return _card_dict(service.activate_rate_card(card_id), include_rows=True)
    except KeyError:
        raise HTTPException(status_code=404, detail="rate card not found")


@app.delete("/api/rate-cards/{card_id}")
def delete_rate_card(card_id: str, admin: User = Depends(require_admin)) -> dict:
    try:
        service.delete_rate_card(card_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="rate card not found")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "deleted", "id": card_id}


class VariablesUpdateRequest(BaseModel):
    """Partial Variables override (D31 tier 1, org-wide defaults) — every field
    optional; only fields actually present in the request body are stored as
    overrides, so unset fields keep falling back to the `variables.yaml` default."""

    real_weight: float | None = None
    opt_weight: float | None = None
    pes_weight: float | None = None
    epic_multiplier: float | None = None
    feature_multiplier: float | None = None
    avg_story_pts: float | None = None
    ai_boost_min: float | None = None
    ai_boost_max: float | None = None
    ba_ratio: float | None = None
    designer_ratio: float | None = None
    devops_ratio: float | None = None
    qa_ratio: float | None = None
    hours_per_sprint: float | None = None
    weeks_in_sprint: int | None = None
    working_month_days: int | None = None
    hours_per_day: float | None = None
    risk_impact_low: float | None = None
    risk_impact_moderate: float | None = None
    risk_impact_high: float | None = None
    risk_impact_extreme: float | None = None
    accelerator_impact_low: float | None = None
    accelerator_impact_moderate: float | None = None
    accelerator_impact_high: float | None = None
    accelerator_impact_extreme: float | None = None
    days_per_story_point: float | None = None


@app.get("/api/variables")
def get_variables(user: User = Depends(get_current_user)) -> dict:
    """Org-wide Variables (D31 tier 1): effective values (YAML defaults with any
    admin overrides applied) plus which fields are currently overridden."""
    return {
        "effective": service.get_variables().model_dump(),
        "overridden_fields": sorted(service.get_variable_overrides().keys()),
    }


@app.put("/api/variables")
def update_variables(req: VariablesUpdateRequest, admin: User = Depends(require_admin)) -> dict:
    """Set org-wide Variables overrides. Only fields present in the request body
    are stored; omit a field to leave it at whatever it currently resolves to
    (default or a previously-set override)."""
    overrides = req.model_dump(exclude_unset=True, exclude_none=True)
    effective = service.update_variables(overrides)
    return {
        "effective": effective.model_dump(),
        "overridden_fields": sorted(service.get_variable_overrides().keys()),
    }


@app.get("/api/tshirt-scale")
def get_tshirt_scale(user: User = Depends(get_current_user)) -> dict:
    """The t-shirt size -> points matrix in effect (D34) — YAML defaults with
    any admin overrides applied. Single-tier, org-wide only (see D31's scoping
    note): unlike Variables, there's no per-estimate version of this."""
    return {
        "sizes": service.get_tshirt_scale(),
        "overridden_sizes": sorted(service.get_tshirt_scale_overrides().keys()),
    }


@app.put("/api/tshirt-scale")
def update_tshirt_scale(overrides: dict[str, dict[str, float]], admin: User = Depends(require_admin)) -> dict:
    """Set org-wide t-shirt scale overrides — size -> {epic/feature/story -> points}.
    Only the sizes/levels present in the request body are stored; the rest keep
    falling back to whatever they currently resolve to."""
    effective = service.update_tshirt_scale(overrides)
    return {
        "sizes": effective,
        "overridden_sizes": sorted(service.get_tshirt_scale_overrides().keys()),
    }


@app.post("/api/estimates/{estimate_id}/recost", response_model=EstimateResponse)
def recost_estimate(estimate_id: str, user: User = Depends(get_current_user)) -> EstimateResponse:
    """Reprice an existing estimate under the active rate card (new version)."""
    _access_or_403(user, estimate_id, "edit")
    try:
        saved = service.recost_estimate(estimate_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="estimate not found")
    return _to_response(saved, service.get_references(estimate_id))


@app.put("/api/estimates/{estimate_id}/context", response_model=EstimateResponse)
def save_context(estimate_id: str, panel: ContextPanel, user: User = Depends(get_current_user)) -> EstimateResponse:
    """Save the Context Panel and auto-recalculate the estimate (in place)."""
    _access_or_403(user, estimate_id, "edit")
    try:
        stored, references = service.recalculate_from_context(estimate_id, panel)
    except KeyError:
        raise HTTPException(status_code=404, detail="estimate not found")
    return _to_response(stored, references)


class UrlIngestRequest(BaseModel):
    url: str


@app.post("/api/ingest/url")
def ingest_url(req: UrlIngestRequest, user: User = Depends(get_current_user)) -> dict:
    """Fetch a URL and return its text for use as a context entry."""
    import httpx

    url = req.url.strip()
    if not (url.startswith("http://") or url.startswith("https://")):
        raise HTTPException(status_code=422, detail="URL must start with http(s)://")
    try:
        resp = httpx.get(url, timeout=15, follow_redirects=True, headers={"User-Agent": "Architect.IQ"})
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=422, detail=f"could not fetch URL: {exc}")
    text = resp.text
    # Strip HTML tags crudely so the model sees readable content.
    if "html" in resp.headers.get("content-type", "").lower() or "<html" in text[:2000].lower():
        text = re.sub(r"<script.*?</script>|<style.*?</style>", " ", text, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
    return {"url": url, "text": text[:20000], "status": "ingested"}


@app.get("/api/ai-tiers")
def list_ai_tiers() -> list[dict]:
    """The AI Tier ladder (human-to-AI-agent ratio) for team modeling."""
    models, _ = data_loader.load_dev_models()
    tiers = [
        {"key": k, "name": m["name"], "tier": m.get("tier", 0),
         "human_ratio": m.get("human_ratio", 1), "ai_ratio": m.get("ai_ratio", 0),
         "human_role": m.get("human_role", ""), "ai_role": m.get("ai_role", ""),
         "ai_boost": m["ai_boost"], "effort_multiplier": m["effort_multiplier"],
         "assumptions": m.get("assumptions", [])}
        for k, m in models.items()
    ]
    return sorted(tiers, key=lambda t: t["tier"])


class ScenariosRequest(BaseModel):
    scenarios: list[Scenario] | None = None


@app.post("/api/estimates/{estimate_id}/scenarios", response_model=EstimateResponse)
def compute_scenarios_endpoint(estimate_id: str, req: ScenariosRequest, user: User = Depends(get_current_user)) -> EstimateResponse:
    """Compute staffing/dev-model scenarios (defaults if none given) and persist."""
    _access_or_403(user, estimate_id, "edit")
    try:
        saved = service.compute_scenarios(estimate_id, req.scenarios)
    except KeyError:
        raise HTTPException(status_code=404, detail="estimate not found")
    return _to_response(saved, service.get_references(estimate_id))


@app.post("/api/estimates/{estimate_id}/suggestions")
def suggestions_endpoint(estimate_id: str, user: User = Depends(get_current_user)) -> dict:
    """Advisor: cheaper/faster team models and scope deferrals (history-grounded)."""
    _access_or_403(user, estimate_id, "view")
    try:
        result = service.suggest(estimate_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="estimate not found")
    return {
        "team": [t.model_dump() for t in result["team"]],
        "deferrals": [d.model_dump() for d in result["deferrals"]],
    }


@app.post("/api/estimates/{estimate_id}/actuals")
def record_actuals(estimate_id: str, outcome: ActualOutcome, user: User = Depends(get_current_user)) -> dict:
    _access_or_403(user, estimate_id, "edit")
    if service.get_estimate(estimate_id) is None:
        raise HTTPException(status_code=404, detail="estimate not found")
    outcome.estimate_id = estimate_id
    service.record_actuals(outcome)
    return {"status": "recorded", "estimate_id": estimate_id}


# --- Sharing, public links, comments ---

class ShareRequest(BaseModel):
    principal: str  # email or known user's name
    permission: Permission = Permission.VIEW


@app.get("/api/estimates/{estimate_id}/access")
def my_access(estimate_id: str, user: User = Depends(get_current_user)) -> dict:
    ctx = resolve_access(user, estimate_id, service.repo, service.directory)
    return ctx.model_dump()


@app.get("/api/estimates/{estimate_id}/shares")
def list_shares(estimate_id: str, user: User = Depends(get_current_user)) -> dict:
    _access_or_403(user, estimate_id, "edit")
    return {
        "shares": [s.model_dump() for s in service.directory.list_shares(estimate_id)],
        "links": [l.model_dump() for l in service.directory.list_share_links(estimate_id)],
    }


@app.post("/api/estimates/{estimate_id}/shares")
def add_share(estimate_id: str, req: ShareRequest, user: User = Depends(get_current_user)) -> dict:
    _access_or_403(user, estimate_id, "edit")
    try:
        email = service.share_principal_email(req.principal)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    service.directory.add_share(estimate_id, email, req.permission)
    return {"status": "shared", "principal_email": email, "permission": req.permission.value}


@app.delete("/api/estimates/{estimate_id}/shares/{email}")
def remove_share(estimate_id: str, email: str, user: User = Depends(get_current_user)) -> dict:
    _access_or_403(user, estimate_id, "edit")
    service.directory.remove_share(estimate_id, email)
    return {"status": "removed", "principal_email": email}


@app.post("/api/estimates/{estimate_id}/share-link")
def create_share_link(estimate_id: str, user: User = Depends(get_current_user)) -> dict:
    _access_or_403(user, estimate_id, "edit")
    link = service.directory.create_share_link(estimate_id, user.email)
    return {"token": link.token, "path": f"/shared/{link.token}"}


@app.delete("/api/estimates/{estimate_id}/share-link/{token}")
def revoke_share_link(estimate_id: str, token: str, user: User = Depends(get_current_user)) -> dict:
    _access_or_403(user, estimate_id, "edit")
    service.directory.revoke_share_link(token)
    return {"status": "revoked", "token": token}


@app.get("/api/shared/{token}", response_model=EstimateResponse)
def shared_estimate(token: str) -> EstimateResponse:
    """Public, view-only access to an estimate via a share-link token (no login)."""
    link = service.directory.get_share_link(token)
    if link is None:
        raise HTTPException(status_code=404, detail="invalid or revoked share link")
    stored = service.get_estimate(link.estimate_id)
    if stored is None:
        raise HTTPException(status_code=404, detail="estimate not found")
    return _to_response(stored, service.get_references(link.estimate_id))


@app.get("/api/estimates/{estimate_id}/comments")
def list_comments(estimate_id: str, user: User = Depends(get_current_user)) -> list[dict]:
    _access_or_403(user, estimate_id, "view")
    return [c.model_dump() for c in service.directory.list_comments(estimate_id)]


class CommentRequest(BaseModel):
    body: str


@app.post("/api/estimates/{estimate_id}/comments")
def add_comment(estimate_id: str, req: CommentRequest, user: User = Depends(get_current_user)) -> dict:
    _access_or_403(user, estimate_id, "comment")
    if not req.body.strip():
        raise HTTPException(status_code=422, detail="empty comment")
    comment = service.directory.add_comment(estimate_id, user.name, req.body.strip())
    return comment.model_dump()


# --- Accounts, opportunities, users (admin manages; authed users read) ---

class AccountRequest(BaseModel):
    name: str
    sf_account_id: str | None = None


class OpportunityRequest(BaseModel):
    name: str
    account_id: str
    sf_opportunity_id: str | None = None
    notion_page_ref: str | None = None


class UserRequest(BaseModel):
    email: str
    name: str
    role: Role = Role.USER
    password: str | None = None


class AssignRequest(BaseModel):
    account_id: str | None = None
    opportunity_id: str | None = None


@app.get("/api/accounts")
def list_accounts(user: User = Depends(get_current_user)) -> list[dict]:
    return [a.model_dump() for a in service.directory.list_accounts()]


@app.post("/api/accounts")
def create_account(req: AccountRequest, admin: User = Depends(require_admin)) -> dict:
    return service.directory.create_account(req.name, req.sf_account_id).model_dump()


@app.get("/api/opportunities")
def list_opportunities(account_id: str | None = None, user: User = Depends(get_current_user)) -> list[dict]:
    opps = service.directory.list_opportunities(account_id)
    if user.role is Role.CLIENT:
        visible = service.directory.visible_opportunity_ids(user.id)
        opps = [o for o in opps if o.id in visible]
    return [o.model_dump() for o in opps]


@app.post("/api/opportunities")
def create_opportunity(req: OpportunityRequest, admin: User = Depends(require_admin)) -> dict:
    return service.directory.create_opportunity(
        name=req.name, account_id=req.account_id,
        sf_opportunity_id=req.sf_opportunity_id, notion_page_ref=req.notion_page_ref,
    ).model_dump()


@app.get("/api/opportunities/{opp_id}")
def get_opportunity(opp_id: str, user: User = Depends(get_current_user)) -> dict:
    opp = service.directory.get_opportunity(opp_id)
    if opp is None:
        raise HTTPException(status_code=404, detail="opportunity not found")
    if user.role is Role.CLIENT and opp_id not in service.directory.visible_opportunity_ids(user.id):
        raise HTTPException(status_code=403, detail="not assigned to this opportunity")
    account = service.directory.get_account(opp.account_id)
    estimates = [s.__dict__ for s in service.list_estimates() if s.opportunity_id == opp_id]
    notes = [n.__dict__ for n in get_opportunity_notes(opp.notion_page_ref)]
    return {"opportunity": opp.model_dump(), "account": account.model_dump() if account else None,
            "estimates": estimates, "notion_notes": notes}


@app.post("/api/opportunities/{opp_id}/active-estimate")
def set_active_estimate(opp_id: str, estimate_id: str, user: User = Depends(get_current_user)) -> dict:
    """Mark one estimate as the opportunity's official one."""
    _access_or_403(user, estimate_id, "edit")
    service.directory.set_active_estimate(opp_id, estimate_id)
    return {"status": "ok", "opportunity_id": opp_id, "active_estimate_id": estimate_id}


@app.get("/api/users")
def list_users(admin: User = Depends(require_admin)) -> list[dict]:
    return [u.model_dump() for u in service.directory.list_users()]


@app.post("/api/users")
def create_user(req: UserRequest, admin: User = Depends(require_admin)) -> dict:
    if service.directory.get_user_by_email(req.email):
        raise HTTPException(status_code=409, detail="a user with that email already exists")
    return service.directory.create_user(
        email=req.email, name=req.name, role=req.role, password=req.password or "changeme",
    ).model_dump()


@app.post("/api/users/{user_id}/role")
def set_user_role(user_id: str, role: Role, admin: User = Depends(require_admin)) -> dict:
    service.directory.set_role(user_id, role)
    return {"status": "ok", "user_id": user_id, "role": role.value}


@app.post("/api/users/{user_id}/assign")
def assign_client(user_id: str, req: AssignRequest, admin: User = Depends(require_admin)) -> dict:
    service.directory.assign_client(user_id, account_id=req.account_id, opportunity_id=req.opportunity_id)
    return {"status": "assigned", "user_id": user_id}
